import os
import openpyxl
import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from tkinter import ttk, messagebox, simpledialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class Employee:
    def __init__(self, name):
        self.name = name

class CensusSystem:
    def __init__(self):
        self.employees = self.load_employees()
        self.census_dir = os.path.join(os.path.expanduser("~"), "Census")
        self.ensure_census_directory()
        self.holidays = self.load_holidays()

    def load_employees(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        employee_file = os.path.join(script_dir, 'Census.txt')
        employees = []
        try:
            with open(employee_file, 'r') as file:
                for line in file:
                    name = line.strip()
                    if name:
                        employees.append(Employee(name))
        except FileNotFoundError:
            print(f"Warning: Census.txt not found in {script_dir}. Using default employee list.")
            employees = [Employee("John Doe"), Employee("Jane Smith"), Employee("Bob Johnson")]
        return employees
    
    def load_holidays(self):
        # TODO: Implement holiday loading logic
        # For now, return an empty list or hardcode some holidays
        return []
    
    def save_record(self, date, name, reason, return_date, details):
        filepath = self.create_weekly_census_file(date)
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook[date.strftime("%A")]
        
        # Find the row for the given name, or use the first empty row
        name_row = None
        for row in range(2, 32):  # Check rows 2 to 31 (30 rows for data)
            if sheet.cell(row=row, column=1).value == name:
                name_row = row
                break
            elif sheet.cell(row=row, column=1).value is None:
                name_row = row
                break
        
        if name_row is None:
            messagebox.showerror("Error", "No available rows in the sheet.")
            return
        
        # Clear existing data for the row
        for col in range(1, 8):
            sheet.cell(row=name_row, column=col).value = None
        
        # Add new data
        sheet.cell(row=name_row, column=1, value=name)
        if reason == "Vacation":
            sheet.cell(row=name_row, column=2, value="X")
            sheet.cell(row=name_row, column=3, value=return_date)
        elif reason == "Scheduled Leave":
            sheet.cell(row=name_row, column=4, value=details)
        elif reason == "Unscheduled Leave":
            sheet.cell(row=name_row, column=5, value="X")
        elif reason == "Day Off":
            sheet.cell(row=name_row, column=6, value="X")
        elif reason == "Other":
            sheet.cell(row=name_row, column=7, value=details)
        
        # Apply formatting
        for col in range(1, 8):
            cell = sheet.cell(row=name_row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        workbook.save(filepath)

    def create_weekly_census_file(self, date):
        week_dates = self.get_week_dates(date)
        filename = f"Census_{week_dates[0].strftime('%m-%d-%Y')}_to_{week_dates[-1].strftime('%m-%d-%Y')}.xlsx"
        filepath = os.path.join(self.census_dir, filename)
        
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # Remove the default sheet
            
            for day_date in week_dates:
                if day_date.strftime("%Y-%m-%d") not in self.holidays:
                    sheet_name = day_date.strftime("%A")  # Full day name
                    sheet = workbook.create_sheet(sheet_name)
                    self.setup_sheet(sheet, day_date)
            
            workbook.save(filepath)
        
        return filepath

    def ensure_census_directory(self):
        if not os.path.exists(self.census_dir):
            os.makedirs(self.census_dir)
    
    def get_week_dates(self, date):
        start = date - timedelta(days=date.weekday())
        return [start + timedelta(days=i) for i in range(5)]

    def setup_sheet(self, sheet, date):
        sheet.title = date.strftime("%A")
        headers = ["Name", "Vacation", "Return Date", "Scheduled Leave", "Unscheduled Leave", "Day Off", "Other"]
        
        # Set column widths
        sheet.column_dimensions['A'].width = 30
        for col in range(2, 8):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Set row height
        sheet.row_dimensions[1].height = 30
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Set default row style
        default_font = Font(name='Calibri', size=11)
        default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        default_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in range(2, 100):  # Assuming up to 100 rows
            for col in range(1, 8):
                cell = sheet.cell(row=row, column=col)
                cell.font = default_font
                cell.alignment = default_alignment
                cell.border = default_border
        
        # Set alternating row colors
        alt_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for row in range(2, 100, 2):
            for col in range(1, 8):
                sheet.cell(row=row, column=col).fill = alt_fill

    def create_new_census_file(self, filepath, date):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Census"
            
        headers = ["Name", "Vacation", "Return Date", "Scheduled Leave", "Unscheduled Leave", "Day Off", "Other"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            sheet.column_dimensions['A'].width = 20
            for col in range(2, 8):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            workbook.save(filepath)

class CensusGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Census")
        self.master.geometry("220x250")  # Adjusted size to accommodate all widgets
        self.census_system = CensusSystem()

        self.create_widgets()

    def create_widgets(self):
        # Date selection
        tk.Label(self.master, text="Select Date:").pack(pady=5)
        self.date_entry = DateEntry(self.master, width=12, background='darkblue', 
                                    foreground='white', borderwidth=2,
                                    date_pattern='mm/dd/yyyy',
                                    state="readonly")
        self.date_entry.pack(pady=5)
        self.date_entry.bind("<<DateEntrySelected>>", self.validate_date)

        # Employee dropdown
        tk.Label(self.master, text="Select Employee:").pack(pady=5)
        self.employee_var = tk.StringVar()
        self.employee_dropdown = ttk.Combobox(self.master, textvariable=self.employee_var)
        self.employee_dropdown['values'] = [employee.name for employee in self.census_system.employees]
        self.employee_dropdown.pack(pady=5)

        # Reason dropdown
        tk.Label(self.master, text="Select Reason:").pack(pady=5)
        self.reason_var = tk.StringVar()
        self.reason_dropdown = ttk.Combobox(self.master, textvariable=self.reason_var)
        self.reason_dropdown['values'] = ["Vacation", "Scheduled Leave", "Unscheduled Leave", "Day Off", "Other"]
        self.reason_dropdown.pack(pady=5)
        self.reason_dropdown.bind("<<ComboboxSelected>>", self.handle_reason_selection)

        # Submit button
        tk.Button(self.master, text="Submit", command=self.submit_record).pack(pady=20)

    def validate_date(self, event):
        selected_date = self.date_entry.get_date()
        if selected_date.weekday() >= 5:  # Saturday or Sunday
            messagebox.showerror("Invalid Date", "Please select a weekday (Monday-Friday).")
            self.date_entry.set_date(self.get_next_weekday(selected_date))

    def get_next_weekday(self, date):
        days_ahead = 0 - date.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        return date + timedelta(days=days_ahead)

    def handle_reason_selection(self, event):
        reason = self.reason_var.get()
        
        # Remove any existing additional widgets
        if hasattr(self, 'additional_frame'):
            self.additional_frame.destroy()

        # Create a new frame for additional inputs
        self.additional_frame = tk.Frame(self.master)
        self.additional_frame.pack(pady=10)

        if reason in ["Scheduled Leave", "Other"]:
            tk.Label(self.additional_frame, text="Details:").pack(side=tk.LEFT)
            self.details_entry = tk.Entry(self.additional_frame, width=20)
            self.details_entry.pack(side=tk.LEFT, padx=5)

    def submit_record(self):
        date = self.date_entry.get_date()
        name = self.employee_var.get()
        reason = self.reason_var.get()
        return_date = ""
        details = ""

        if not all([date, name, reason]):
            messagebox.showerror("Error", "Please fill in all required fields.")
            return

        if reason == "Vacation":
            return_date = simpledialog.askstring("Vacation Return Date", 
                                                 "Enter the return date (MM/DD/YYYY):",
                                                 parent=self.master)
            if not return_date:
                messagebox.showerror("Error", "Return date is required for vacation.")
                return
            try:
                # Validate the date format
                datetime.strptime(return_date, "%m/%d/%Y")
            except ValueError:
                messagebox.showerror("Error", "Invalid date format. Please use MM/DD/YYYY.")
                return
        elif reason in ["Scheduled Leave", "Other"]:
            if hasattr(self, 'details_entry'):
                details = self.details_entry.get()
            if not details:
                details = simpledialog.askstring(f"{reason} Details", f"Please provide details for '{reason}':")
                if not details:
                    messagebox.showerror("Error", f"Details are required for '{reason}'.")
                    return

        self.census_system.save_record(date, name, reason, return_date, details)
        messagebox.showinfo("Success", "Record saved successfully.")
        self.reset_form()

    def reset_form(self):
        self.date_entry.set_date(datetime.now())
        self.employee_var.set('')
        self.reason_var.set('')
        if hasattr(self, 'additional_frame'):
            self.additional_frame.destroy()

def main():
    root = tk.Tk()
    CensusGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()